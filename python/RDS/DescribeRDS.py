# -*- coding: utf-8 -*-
"""
@Time ： 2021/10/13 11:46
@Auth ： ndmiao
@Blog ：www.ndmiao.cn
"""

import openpyxl
import boto3
from ReadCredentials import ReadCredentials
from openpyxl.styles import Alignment


class GetRDSInformation:
    def __init__(self):
        self.region = 'cn-north-1'

    def rds_client(self, id, key, region):
        """
        :param id: aws_access_key_id
        :param key: aws_secret_access_key
        :param region: region_name
        :return: 建立一个与 rds 的连接
        """
        rds = boto3.client(
            'rds',
            aws_access_key_id = id,
            aws_secret_access_key = key,
            region_name = region
        )
        return rds

    def rds_information(self):
        """
        :return: 获取所有账号下rds资源信息，保存到xlsx里面
        """
        credentials = ReadCredentials().get_credential()
        wb = WriteToXlsx().active_xlsx()
        id_num = 0
        for credential in credentials:
            rds = self.rds_client(credential['Access key ID'], credential['Secret access key'], self.region)
            rds_instances = rds.describe_db_instances()
            instance_info = []
            owner_id = ReadCredentials().aws_id()[id_num]
            id_num += 1
            for item in rds_instances['DBInstances']:
                Name = item['DBInstanceIdentifier']
                InstanceType = item['DBInstanceClass']
                Engine = item['Engine']
                EngineVersion = item['EngineVersion']
                State = item['DBInstanceStatus']
                Tags = item['TagList']
                Project = self.get_tag(Tags, 'project')
                Schedule = self.get_tag(Tags, 'Schedule')
                ScheduleMessage = self.get_tag(Tags, 'ScheduleMessage')
                instance_info.append([Name, InstanceType, Engine, EngineVersion, State, Project, Schedule, ScheduleMessage])
            WriteToXlsx().send_data(wb, owner_id, instance_info)
        WriteToXlsx().save(wb)

    def get_tag(self, Tags, key):
        value = [tag['Value'] for tag in Tags if tag['Key'] == key]
        try:
            value = value[0]
        except:
            value = 'None'
        return value

class WriteToXlsx:
    def __init__(self):
        self.headers = ['Name', 'InstanceType', 'Engine', 'EngineVersion', 'State', 'Project', 'Schedule', 'ScheduleMessage']
        self.xlsx_name = 'D:/BaiduNetdiskWorkspace/代码/python/aws/RDS/ResoursesList.xlsx'

    def active_xlsx(self):
        """
        :return: 创建一个 xlsx 连接
        """
        wb = openpyxl.Workbook()
        return wb

    def send_data(self, wb, sheet_name, sheet_data):
        """
        :param wb: wb
        :param sheet_name: 表名
        :param sheet_data: 写入的数据
        :return: 生成一个 sheet 并写入数据
        """
        ws = wb.create_sheet(sheet_name)
        ws.append(self.headers)
        for data in sheet_data:
            ws.append(data)
        self.center(wb, sheet_name)

    def center(self, wb, sheet_name):
        table = wb[sheet_name]
        rows = table.max_row
        cols = table.max_column
        alignobj = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for r in range(rows + 1):
            for c in range(cols + 1):
                if r != 0 and c != 0:
                    table.cell(row=r, column=c).alignment = alignobj

    def save(self, wb):
        """
        :return: 保存
        """
        wb.remove(wb['Sheet'])
        wb.save(self.xlsx_name)


if __name__ == "__main__":
    # wb = WriteToXlsx().active_xlsx()
    # WriteToXlsx().send_data(wb, 'hhhh', [[1,2,3],[4,5,6]])
    # WriteToXlsx().send_data(wb, 'hhhh3', [[1, 2, 3], [4, 5, 6 ]])
    # WriteToXlsx().save(wb)
    GetRDSInformation().rds_information()

