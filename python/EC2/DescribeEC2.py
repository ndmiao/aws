# -*- coding: utf-8 -*-
"""
@Time ： 2021/10/9 10:07
@Auth ： ndmiao
@Blog ：www.ndmiao.cn
@Url ：https://boto3.amazonaws.com/v1/documentation/api/latest/reference/services/ec2.html#EC2.Client.describe_instances
"""

import openpyxl
import boto3
from ReadCredentials import ReadCredentials
from openpyxl.styles import Alignment


class GetEC2Information:
    def __init__(self):
        self.region = 'cn-north-1'

    def ec2_client(self, id, key, region):
        """
        :param id: aws_access_key_id
        :param key: aws_secret_access_key
        :param region: region_name
        :return: 建立一个与 ec2 的连接
        """
        ec2 = boto3.client(
            'ec2',
            aws_access_key_id = id,
            aws_secret_access_key = key,
            region_name = region
        )
        return ec2

    def ec2_information(self):
        """
        :return: 获取所有账号下ec2资源信息，保存到xlsx里面
        """
        credentials = ReadCredentials().get_credential()
        wb = WriteToXlsx().active_xlsx()
        for credential in credentials:
            ec2 = self.ec2_client(credential['Access key ID'], credential['Secret access key'], self.region)
            instances = ec2.describe_instances()
            instance_info = []
            for item in instances['Reservations']:
                owner_id = item['OwnerId']
                for instance in item['Instances']:
                    InstanceId = instance['InstanceId']
                    InstanceType = instance['InstanceType']
                    PrivateIpAddress = instance['PrivateIpAddress']
                    State = instance['State']['Name']
                    Tags = instance['Tags']
                    Name = self.get_tag(Tags, 'Name')
                    Project = self.get_tag(Tags, 'project')
                    Schedule = self.get_tag(Tags, 'Schedule')
                    ScheduleMessage = self.get_tag(Tags, 'ScheduleMessage')
                    instance_info.append([Name, InstanceId, InstanceType, PrivateIpAddress, Project, State, Schedule, ScheduleMessage])
            WriteToXlsx().send_data(wb, owner_id, instance_info)
        WriteToXlsx().save(wb)

    def get_tag(self, Tags, key):
        value = [tag['Value'] for tag in Tags if tag['Key'] == key]
        try:
            value = value[0]
        except:
            value = 'None'
        return value

class WriteToXlsx:
    def __init__(self):
        self.headers = ['Name', 'InstanceId', 'InstanceType', 'PrivateIpAddress', 'Project', 'State', 'Schedule', 'ScheduleMessage']
        self.xlsx_name = 'D:/BaiduNetdiskWorkspace/代码/python/aws/EC2/ResoursesList.xlsx'

    def active_xlsx(self):
        """
        :return: 创建一个 xlsx 连接
        """
        wb = openpyxl.Workbook()
        return wb

    def send_data(self, wb, sheet_name, sheet_data):
        """
        :param wb: wb
        :param sheet_name: 表名
        :param sheet_data: 写入的数据
        :return: 生成一个 sheet 并写入数据
        """
        ws = wb.create_sheet(sheet_name)
        ws.append(self.headers)
        for data in sheet_data:
            ws.append(data)
        self.center(wb, sheet_name)

    def center(self, wb, sheet_name):
        table = wb[sheet_name]
        rows = table.max_row
        cols = table.max_column
        alignobj = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for r in range(rows + 1):
            for c in range(cols + 1):
                if r != 0 and c != 0:
                    table.cell(row=r, column=c).alignment = alignobj

    def save(self, wb):
        """
        :return: 保存
        """
        wb.remove(wb['Sheet'])
        wb.save(self.xlsx_name)


if __name__ == "__main__":
    # wb = WriteToXlsx().active_xlsx()
    # WriteToXlsx().send_data(wb, 'hhhh', [[1,2,3],[4,5,6]])
    # WriteToXlsx().send_data(wb, 'hhhh3', [[1, 2, 3], [4, 5, 6 ]])
    # WriteToXlsx().save(wb)
    GetEC2Information().ec2_information()